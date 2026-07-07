from __future__ import annotations

import json
from pathlib import Path

import cv2
import numpy as np
from openpyxl import load_workbook

from cad_dimensions.geometry import BBox
from cad_dimensions.models import ExtractionResult, dimension_to_record
from cad_dimensions.models import DimensionRow
from cad_dimensions.pipeline import write_result_artifacts
from decimal import Decimal


def test_write_structured_artifacts_and_crop(tmp_path: Path) -> None:
    image_path = tmp_path / "page.png"
    image = np.full((120, 200, 3), 255, dtype=np.uint8)
    cv2.putText(image, "12 +/-1", (20, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1)
    cv2.imwrite(str(image_path), image)

    row = DimensionRow(
        "A1",
        Decimal("12"),
        "+/-1",
        Decimal("1"),
        Decimal("1"),
        Decimal("1"),
        source="ocr",
        source_bbox=BBox(15, 35, 100, 70),
        status="review",
    )
    record = dimension_to_record(row, "4022.000.0000", "sample.pdf", index=1)
    result = ExtractionResult(Path("sample.pdf"), "4022.000.0000", "generic-ocr-review", [record], 0, {"notes": "test"})

    artifacts = write_result_artifacts(result, tmp_path / "run", image_path)

    dimensions = json.loads(artifacts["dimensions_json"].read_text(encoding="utf-8"))
    assert dimensions["dimensions"][0]["status"] == "review"
    assert Path(dimensions["dimensions"][0]["crop_path"]).exists()

    workbook = load_workbook(artifacts["excel"], read_only=True)
    assert workbook["Dimensions"].max_row == 2
