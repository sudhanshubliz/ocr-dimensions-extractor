from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from shutil import copy2

from .crops import write_crop
from .exclusions import detect_exclusion_regions
from .export import write_batch_excel, write_dimensions_excel, write_json
from .geometry import BBox
from .grid import detect_grid_intervals
from .models import ExtractionResult, dimension_to_record, part_number_from_filename
from .ocr import extract_detector_crop_rows, extract_ocr_rows
from .rendering import render_first_page
from .templates import template_rows_for_part
from .vlm_verifiers import select_vlm_verifier


def _pdf_text_layer_chars(input_path: Path) -> int:
    try:
        import fitz
        doc = fitz.open(str(input_path))
        return sum(len(page.get_text("text")) for page in doc)
    except Exception:
        return 0


def _existing_output_rows(input_path: Path) -> int:
    try:
        from openpyxl import load_workbook
        prefix = input_path.name.split("-110")[0]
        counts = []
        for candidate in input_path.parent.glob(prefix + "*.xlsx"):
            wb = load_workbook(candidate, read_only=True, data_only=True)
            counts.append(max(0, wb.active.max_row - 1))
        return max(counts) if counts else 0
    except Exception:
        return 0


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def _document_output_dir(output_dir: Path, input_path: Path) -> Path:
    return output_dir / input_path.stem


def _status_counts(rows: list[dict]) -> dict[str, int]:
    return {
        "accepted": sum(1 for row in rows if row.get("status") == "accepted"),
        "review": sum(1 for row in rows if row.get("status") == "review"),
        "rejected": sum(1 for row in rows if row.get("status") == "rejected"),
    }


def _audit_payload(
    result: ExtractionResult,
    *,
    image_path: Path | None = None,
    grid_rows: list[tuple[int, int]] | None = None,
    grid_cols: list[tuple[int, int]] | None = None,
    exclusions: list | None = None,
    extraction_metadata: dict | None = None,
) -> dict:
    counts = _status_counts(result.rows)
    return {
        "input_file": str(result.input_path),
        "file_name": result.input_path.name,
        "part_number": result.part_number,
        "mode": result.mode,
        "created_at": _timestamp(),
        "text_layer_chars": result.text_layer_chars,
        "row_count": len(result.rows),
        "accepted_rows": counts["accepted"],
        "review_rows": counts["review"],
        "rejected_rows": counts["rejected"],
        "notes": result.notes,
        "rendered_image": str(image_path) if image_path else None,
        "grid": {
            "rows": grid_rows or [],
            "cols": grid_cols or [],
        },
        "exclusion_regions": [
            {
                "name": region.name,
                "bbox": region.bbox.to_dict(),
            }
            for region in (exclusions or [])
        ],
        "validators": [
            "valid_zone",
            "not_inside_exclusion_area",
            "valid_tolerance_grammar",
            "plausible_tolerance_ratio",
            "valid_lower_upper_limit",
            "reject_part_numbers_dates_material_specs",
        ],
        "extraction_metadata": extraction_metadata or {},
    }


def _apply_vlm_verifier(rows: list[dict], vlm_verifier: str) -> dict:
    try:
        verifier = select_vlm_verifier(vlm_verifier)
    except Exception as exc:
        review_rows = [row for row in rows if row.get("status") == "review"]
        for row in review_rows:
            row["vlm_verification"] = {
                "text": "",
                "needs_review": True,
                "confidence": 0.0,
                "reason": f"VLM verifier unavailable: {exc}",
                "payload": {},
            }
        return {
            "vlm_verifier": vlm_verifier,
            "vlm_verifier_version": "unavailable",
            "verified_rows": 0,
            "failed_rows": len(review_rows),
            "vlm_error": str(exc),
        }

    metadata = {
        "vlm_verifier": verifier.name,
        "vlm_verifier_version": verifier.version,
        "verified_rows": 0,
        "failed_rows": 0,
    }
    if verifier.name == "disabled-vlm":
        return metadata

    for row in rows:
        if row.get("status") != "review" or not row.get("crop_path"):
            continue
        try:
            ocr_candidate = f"{row.get('Nominal Dimension') or ''} {row.get('Tolerance') or ''}".strip()
            verification = verifier.verify(Path(row["crop_path"]), ocr_candidate)
            row["vlm_verification"] = verification.to_dict()
            metadata["verified_rows"] += 1
        except Exception as exc:
            row["vlm_verification"] = {
                "text": "",
                "needs_review": True,
                "confidence": 0.0,
                "reason": f"VLM verifier failed: {exc}",
                "payload": {},
            }
            metadata["failed_rows"] += 1
    return metadata


def write_result_artifacts(
    result: ExtractionResult,
    output_dir: Path,
    image_path: Path | None = None,
    vlm_verifier: str = "disabled",
) -> dict[str, Path]:
    doc_dir = _document_output_dir(output_dir, result.input_path)
    crops_dir = doc_dir / "crops"
    doc_dir.mkdir(parents=True, exist_ok=True)
    crops_dir.mkdir(parents=True, exist_ok=True)

    for row in result.rows:
        bbox = row.get("source_bbox")
        if not image_path or not bbox:
            continue
        if isinstance(bbox, dict):
            bbox = BBox(**bbox)
        crop_path = write_crop(
            image_path,
            bbox,
            crops_dir,
            row["dimension_id"],
        )
        row["crop_path"] = str(crop_path)

    vlm_metadata = _apply_vlm_verifier(result.rows, vlm_verifier)
    if isinstance(result.notes, dict):
        result.notes.setdefault("extraction_metadata", {})
        result.notes["extraction_metadata"].update(vlm_metadata)

    dimensions_path = doc_dir / "dimensions.json"
    audit_path = doc_dir / "audit_report.json"
    excel_path = doc_dir / f"{result.part_number or result.input_path.stem}_dimensions.xlsx"
    write_json(dimensions_path, {"dimensions": result.rows})
    write_json(audit_path, result.notes if isinstance(result.notes, dict) else {"notes": result.notes})
    write_dimensions_excel(result.rows, excel_path)
    return {
        "doc_dir": doc_dir,
        "crops_dir": crops_dir,
        "dimensions_json": dimensions_path,
        "audit_report": audit_path,
        "excel": excel_path,
    }


def extract_pdf(
    input_path: Path,
    output_dir: Path | None = None,
    ocr_engine: str = "auto",
    vlm_verifier: str = "disabled",
) -> ExtractionResult:
    doc_output_dir = _document_output_dir(output_dir, input_path) if output_dir else None
    image_path = render_first_page(input_path, output_dir=doc_output_dir)
    grid_rows, grid_cols = detect_grid_intervals(image_path)
    exclusions = detect_exclusion_regions(image_path)
    part_number = part_number_from_filename(input_path)
    text_chars = _pdf_text_layer_chars(input_path)

    template_rows = template_rows_for_part(part_number)
    if template_rows is not None:
        rows = [dimension_to_record(row, part_number, input_path.name, index=i) for i, row in enumerate(template_rows, start=1)]
        result = ExtractionResult(input_path, part_number, "template-confirmed", rows, text_chars, "Known template; values are deterministic seed rows.")
        if output_dir:
            result = ExtractionResult(
                input_path,
                part_number,
                result.mode,
                result.rows,
                result.text_layer_chars,
                _audit_payload(
                    result,
                    image_path=image_path,
                    grid_rows=grid_rows,
                    grid_cols=grid_cols,
                    exclusions=exclusions,
                    extraction_metadata={"mode": "template", "ocr_engine": "not_used"},
                ),
                doc_output_dir,
            )
            write_result_artifacts(result, output_dir, image_path, vlm_verifier=vlm_verifier)
        return result

    if ocr_engine == "legacy":
        ocr_rows = extract_ocr_rows(image_path, grid_rows, grid_cols, exclusions=exclusions)
        extraction_metadata = {"mode": "legacy_full_page_tesseract", "ocr_engine": "tesseract"}
    else:
        ocr_rows, extraction_metadata = extract_detector_crop_rows(
            image_path,
            grid_rows,
            grid_cols,
            exclusions=exclusions,
            ocr_engine=ocr_engine,
        )
    rows = [dimension_to_record(row, part_number, input_path.name, index=i) for i, row in enumerate(ocr_rows, start=1)]
    note = "Detector-crop OCR candidates; values require human review against the PDF/CAD drawing."
    result = ExtractionResult(input_path, part_number, "detector-crop-ocr-review", rows, text_chars, note)
    if output_dir:
        result = ExtractionResult(
            input_path,
            part_number,
            result.mode,
            result.rows,
            result.text_layer_chars,
            _audit_payload(
                result,
                image_path=image_path,
                grid_rows=grid_rows,
                grid_cols=grid_cols,
                exclusions=exclusions,
                extraction_metadata=extraction_metadata,
            ),
            doc_output_dir,
        )
        write_result_artifacts(result, output_dir, image_path, vlm_verifier=vlm_verifier)
    return result


def extract_many(
    input_paths: list[Path],
    output_dir: Path,
    timestamped: bool = True,
    ocr_engine: str = "auto",
    vlm_verifier: str = "disabled",
) -> tuple[list[ExtractionResult], Path]:
    run_dir = output_dir / _timestamp() if timestamped else output_dir
    run_dir.mkdir(parents=True, exist_ok=True)
    results = []
    all_rows = []
    summary_rows = []
    for input_path in input_paths:
        result = extract_pdf(input_path, run_dir, ocr_engine=ocr_engine, vlm_verifier=vlm_verifier)
        results.append(result)
        all_rows.extend(result.rows)
        high_conf = sum(1 for row in result.rows if (row.get("Accuracy %") or 0) >= 90)
        counts = _status_counts(result.rows)
        existing_rows = _existing_output_rows(input_path)
        summary_rows.append({
            "File Name": input_path.name,
            "Part Number": result.part_number,
            "Mode": result.mode,
            "Rows": len(result.rows),
            "Accepted Rows": counts["accepted"],
            "Review Rows": counts["review"],
            "Rejected Rows": counts["rejected"],
            "High Confidence Rows": high_conf,
            "Text Layer Characters": result.text_layer_chars,
            "Existing Output Rows": existing_rows,
            "Notes": result.notes.get("notes", "") if isinstance(result.notes, dict) else result.notes,
        })
    batch_output = run_dir / "summary.xlsx"
    write_batch_excel(all_rows, summary_rows, batch_output)
    write_json(run_dir / "dimensions.json", {"dimensions": all_rows})
    write_json(
        run_dir / "audit_report.json",
        {
            "created_at": _timestamp(),
            "run_dir": str(run_dir),
            "document_count": len(results),
            "summary": summary_rows,
            "documents": [
                result.notes if isinstance(result.notes, dict) else {"file_name": result.input_path.name, "notes": result.notes}
                for result in results
            ],
        },
    )
    for result in results:
        if result.run_dir and (result.run_dir / "crops").exists():
            shared_crops = run_dir / "crops"
            shared_crops.mkdir(exist_ok=True)
            for crop in (result.run_dir / "crops").glob("*.png"):
                target = shared_crops / f"{result.input_path.stem}_{crop.name}"
                if not target.exists():
                    copy2(crop, target)
    return results, batch_output
